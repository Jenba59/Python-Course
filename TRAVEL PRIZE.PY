import random
import pycountry
import openpyxl

participants = []
amount = []
country = []

with open ('LISTA24.txt', 'r') as f:
    for line in f:
        participants.append(line.strip())

initialamount = random.randint(1,100000)
for i in range(len(participants)):
    randomamount = random.randint (0,100000)
    initialamount += randomamount
    amount.append(randomamount)

for i in range (len(participants)):
    randomnum = random.randint(0,244)
    country.append(list(pycountry.countries)[randomnum].name)

wb = openpyxl.Workbook()
sheet = wb.active

sheet.cell(row=1, column=1).value = "NAME"
sheet.cell(row=1, column=2).value = "AMOUNT"
sheet.cell(row=1, column=3).value = "COUNTRY"

for i, j, k in zip(participants, amount, country):
    sheet.cell(row=participants.index(i) + 2, column=1).value = i
    sheet.cell(row=participants.index(i) + 2, column=2).value = j
    sheet.cell(row=participants.index(i) + 2, column=3).value = k

sheet.cell(row=len(participants)+2, column=2).value = initialamount

wb.save("SORTEO.xlsx")
